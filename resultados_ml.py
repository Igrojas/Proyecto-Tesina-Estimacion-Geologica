# %%
"""
Visualización y análisis de resultados de predicción (KNN y XGBoost).
Carga el DataFrame df_relleno generado por train_ml.py y genera gráficos.
"""

import warnings
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd

warnings.filterwarnings("ignore")

# --- Configuración ---
INPUT_DF_RELLENO_PATH = Path("data/processed/df_relleno.csv")
INPUT_COMBINED_PATH = Path("data/processed/cluster/puntos_originales_y_relleno.csv")
OUTPUT_TABLA_TESINA_PATH = Path("data/processed/tabla_resultados_ml_tesina.xlsx")
IMAGENES_DIR = Path("imagenes")
COORD_COLS = ["Este", "Norte", "Cota"]
COORD_LABELS = ["Este (m)", "Norte (m)", "Cota (m)"]
ORIGEN_COL = "origen"
VALUE_COL_ORIG = "Rec_Peso_PND25_(%)_nscore"
VALUE_LABEL = "Recuperación en peso (%) (normal score)"


def setup_report_style() -> None:
    """Estilo de figuras (misma línea que Analisis_EDA / Analisis_cluster)."""
    plt.rcParams.update({
        "figure.dpi": 150,
        "savefig.dpi": 300,
        "savefig.bbox": "tight",
        "font.family": "sans-serif",
        "font.size": 11,
        "axes.titlesize": 12,
        "axes.labelsize": 11,
        "xtick.labelsize": 10,
        "ytick.labelsize": 10,
        "axes.grid": True,
        "grid.alpha": 0.35,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "figure.facecolor": "white",
        "axes.facecolor": "white",
    })


def drift_by_bins(
    df: pd.DataFrame,
    coord_col: str,
    value_col: str,
    n_bins: int = 10,
) -> pd.DataFrame:
    """Promedio de value_col por intervalos (bins) de coord_col."""
    out = df[[coord_col, value_col]].copy()
    out["bin"] = pd.cut(out[coord_col], bins=n_bins)
    agg = out.groupby("bin", observed=True)[value_col].agg(["mean", "count"])
    agg["coord_center"] = agg.index.map(lambda x: x.mid)
    return agg.reset_index()


def plot_drift_comparison(
    df_orig: pd.DataFrame,
    df_relleno: pd.DataFrame,
    coord_cols: list[str],
    n_bins: int = 20,
    coord_labels: list[str] | None = None,
    value_label: str | None = None,
    save_path: Path | None = None,
) -> None:
    """Deriva (promedio por bins) para cada coordenada: Original, KNN y XGBoost en la misma gráfica."""
    x_labels = coord_labels if coord_labels is not None else coord_cols
    y_label = value_label or VALUE_LABEL
    n = len(coord_cols)
    fig, axes = plt.subplots(1, n, figsize=(5 * n, 4))
    if n == 1:
        axes = [axes]
    for ax, coord, x_label in zip(axes, coord_cols, x_labels):
        dr_orig = drift_by_bins(df_orig, coord, VALUE_COL_ORIG, n_bins=n_bins)
        dr_knn = drift_by_bins(df_relleno, coord, "pred_nscore_knn", n_bins=n_bins)
        dr_xgb = drift_by_bins(df_relleno, coord, "pred_nscore_xgb", n_bins=n_bins)
        ax.plot(dr_orig["coord_center"], dr_orig["mean"], "o-", linewidth=1, markersize=5, color="#7c3aed", label="Original")
        ax.plot(dr_knn["coord_center"], dr_knn["mean"], "o--", linewidth=1, markersize=5, color="#2563eb", label="KNN")
        ax.plot(dr_xgb["coord_center"], dr_xgb["mean"], "o--", linewidth=1, markersize=5, color="#059669", label="XGBoost")
        ax.set_xlabel(x_label)
        ax.set_ylabel(f"Promedio {y_label}")
        ax.set_title(f"Deriva según {x_label}")
        ax.legend()
        ax.grid(True, alpha=0.35)
    plt.tight_layout()
    if save_path:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        plt.savefig(save_path)
    plt.show()


def plot_3d_cmap(
    df: pd.DataFrame,
    coord_cols: list[str],
    color_col: str,
    coord_labels: list[str],
    title: str,
    cmap: str = "jet",
    save_path: Path | None = None,
) -> None:
    """Scatter 3D coloreado por una variable continua (cmap)."""
    from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

    fig = plt.figure(figsize=(8, 6))
    ax = fig.add_subplot(111, projection="3d")
    sc = ax.scatter(
        df[coord_cols[0]], df[coord_cols[1]], df[coord_cols[2]],
        c=df[color_col], cmap=cmap, s=15, alpha=0.7,
    )
    ax.set_xlabel(coord_labels[0])
    ax.set_ylabel(coord_labels[1])
    ax.set_zlabel(coord_labels[2])
    ax.set_title(title)
    plt.colorbar(sc, ax=ax, shrink=0.5, pad=0.12, label=color_col)
    plt.tight_layout()
    if save_path:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        plt.savefig(save_path)
    plt.show()


def export_tabla_resultados_tesina(
    df_orig: pd.DataFrame,
    df_relleno: pd.DataFrame,
    output_path: Path,
    tabla_num: int = 1,
    titulo_variable: str | None = None,
) -> None:
    """Genera un Excel en formato de tabla autoexplicativa para la tesina.

    Incluye caption, encabezados con numeración (1)-(6), filas para Datos originales,
    KNN y XGBoost (Promedio, Desv. Est., Mín, Máx., Diferencia de medias) y nota al pie.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        raise ImportError("Se requiere openpyxl para exportar la tabla. Instala con: pip install openpyxl")

    col_orig = VALUE_COL_ORIG
    mean_orig = df_orig[col_orig].mean()
    std_orig = df_orig[col_orig].std()
    min_orig = df_orig[col_orig].min()
    max_orig = df_orig[col_orig].max()
    n_orig = len(df_orig)

    mean_knn = df_relleno["pred_nscore_knn"].mean()
    std_knn = df_relleno["pred_nscore_knn"].std()
    min_knn = df_relleno["pred_nscore_knn"].min()
    max_knn = df_relleno["pred_nscore_knn"].max()
    n_relleno = len(df_relleno)

    mean_xgb = df_relleno["pred_nscore_xgb"].mean()
    std_xgb = df_relleno["pred_nscore_xgb"].std()
    min_xgb = df_relleno["pred_nscore_xgb"].min()
    max_xgb = df_relleno["pred_nscore_xgb"].max()

    diff_knn = mean_knn - mean_orig
    diff_xgb = mean_xgb - mean_orig

    titulo = titulo_variable or (
        "Recuperación en peso (%) (normal score): datos originales y predicciones KNN y XGBoost"
    )
    caption = (
        f"Tabla {tabla_num}: Estadísticas descriptivas de {titulo}, "
        f"n (original) = {n_orig}, n (relleno) = {n_relleno}. "
        "(Letra 10, espaciado simple.)"
    )
    nota = (
        "(a) Elaboración propia en base a datos procesados (cluster con nscore y predicciones de modelos KNN y XGBoost). "
        "(b) n = tamaño de muestra. "
        "Desv. Est. = desviación estándar. "
        "Mín = mínimo valor observado. "
        "Máx. = máximo valor observado. "
        "Diferencia de medias = diferencia respecto a la media de datos originales."
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados ML"

    font_caption = Font(size=10)
    font_header = Font(bold=True, size=10)
    font_cell = Font(size=10)
    align_wrap = Alignment(wrap_text=True, vertical="top")

    # Fila 1: caption
    ws["A1"] = caption
    ws["A1"].font = font_caption
    ws["A1"].alignment = align_wrap
    ws.merge_cells("A1:F1")

    # Fila 3: encabezados principales
    headers = ["Muestra", "Promedio", "Desv. Est.", "Mín", "Máx.", "Diferencia de medias"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = font_header

    # Fila 4: numeración (1)-(6)
    for c in range(1, 7):
        ws.cell(row=4, column=c, value=f"({c})").font = font_cell

    # Filas de datos
    def fmt_num(x: float) -> str:
        return f"{x:.4f}" if pd.notna(x) else "—"

    data_rows = [
        ("Datos originales", mean_orig, std_orig, min_orig, max_orig, "—"),
        ("KNN", mean_knn, std_knn, min_knn, max_knn, diff_knn),
        ("XGBoost", mean_xgb, std_xgb, min_xgb, max_xgb, diff_xgb),
    ]
    for i, row in enumerate(data_rows, start=5):
        muestra, prom, desv, mn, mx, diff = row
        ws.cell(row=i, column=1, value=muestra).font = font_cell
        ws.cell(row=i, column=2, value=fmt_num(prom) if isinstance(prom, (int, float)) else prom).font = font_cell
        ws.cell(row=i, column=3, value=fmt_num(desv) if isinstance(desv, (int, float)) else desv).font = font_cell
        ws.cell(row=i, column=4, value=fmt_num(mn) if isinstance(mn, (int, float)) else mn).font = font_cell
        ws.cell(row=i, column=5, value=fmt_num(mx) if isinstance(mx, (int, float)) else mx).font = font_cell
        if isinstance(diff, str):
            ws.cell(row=i, column=6, value=diff).font = font_cell
        else:
            ws.cell(row=i, column=6, value=fmt_num(diff)).font = font_cell

    # Nota al pie
    row_nota = 9
    ws.cell(row=row_nota, column=1, value="Nota:").font = font_header
    ws.cell(row=row_nota + 1, column=1, value=nota).font = font_caption
    ws.cell(row=row_nota + 1, column=1).alignment = align_wrap
    ws.merge_cells(f"A{row_nota + 1}:F{row_nota + 1}")

    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Tabla para tesina guardada en: {output_path}")


# --- Carga y visualización ---
# %%
if __name__ == "__main__":
    setup_report_style()
    IMAGENES_DIR.mkdir(parents=True, exist_ok=True)

    df_relleno = pd.read_csv(INPUT_DF_RELLENO_PATH)
    df_full = pd.read_csv(INPUT_COMBINED_PATH)
    df_orig = df_full[df_full[ORIGEN_COL] == "original"].copy()

    if df_relleno.empty:
        raise FileNotFoundError(
            f"No hay datos en {INPUT_DF_RELLENO_PATH}. "
            "Ejecuta antes train_ml.py para generar el DataFrame con predicciones."
        )

    # --- Tabla para tesina (formato autoexplicativo) ---
    export_tabla_resultados_tesina(
        df_orig,
        df_relleno,
        OUTPUT_TABLA_TESINA_PATH,
        tabla_num=1,
    )

    # --- Scatter 3D por predicción ---
    plot_3d_cmap(
        df_relleno, COORD_COLS, "pred_nscore_knn", COORD_LABELS, "Predicción KNN",
        save_path=IMAGENES_DIR / "resultados_ml_3d_prediccion_KNN.png",
    )
    plot_3d_cmap(
        df_relleno, COORD_COLS, "pred_nscore_xgb", COORD_LABELS, "Predicción XGBoost",
        save_path=IMAGENES_DIR / "resultados_ml_3d_prediccion_XGBoost.png",
    )

    # --- Histograma: distribución de predicciones vs original ---
    plt.figure(figsize=(8, 5))
    plt.hist(df_relleno["pred_nscore_knn"], bins="auto", alpha=0.7, color="#2563eb", edgecolor="white", label="KNN", density=True)
    plt.hist(df_relleno["pred_nscore_xgb"], bins="auto", alpha=0.7, color="#059669", edgecolor="white", label="XGBoost", density=True)
    plt.hist(df_orig["Rec_Peso_PND25_(%)_nscore"], bins="auto", alpha=0.7, color="#7c3aed", edgecolor="white", label="Original", density=True)
    plt.xlabel("Predicción")
    plt.ylabel("Frecuencia")
    plt.title("Distribución de predicciones")
    plt.legend()
    plt.grid(True, alpha=0.35, axis="y")
    plt.tight_layout()
    plt.savefig(IMAGENES_DIR / "resultados_ml_histograma_distribucion.png")
    plt.show()

    # --- Deriva: Original vs KNN vs XGBoost (misma gráfica por coordenada) ---
    plot_drift_comparison(
        df_orig,
        df_relleno,
        COORD_COLS,
        n_bins=20,
        coord_labels=COORD_LABELS,
        value_label=VALUE_LABEL,
        save_path=IMAGENES_DIR / "resultados_ml_deriva_comparacion.png",
    )

    # --- Agregados por cluster ---
    agg_pred_por_cluster = df_relleno.groupby("cluster_con_nscore")[["pred_nscore_knn", "pred_nscore_xgb"]].agg(
        ["mean", "min", "max"]
    )
    agg_real_por_cluster = df_orig.groupby("cluster_con_nscore")["Rec_Peso_PND25_(%)_nscore"].agg(
        ["mean", "min", "max"]
    )

    # Media por cluster
    plt.figure(figsize=(8, 5))
    plt.plot(
        agg_pred_por_cluster.index,
        agg_pred_por_cluster["pred_nscore_knn"]["mean"],
        marker="o",
        label="KNN (relleno)",
    )
    plt.plot(
        agg_pred_por_cluster.index,
        agg_pred_por_cluster["pred_nscore_xgb"]["mean"],
        marker="o",
        label="XGBoost (relleno)",
    )
    plt.plot(
        agg_real_por_cluster.index,
        agg_real_por_cluster["mean"],
        marker="o",
        label="Original",
        linestyle="--",
        color="#7c3aed",
    )
    plt.xlabel("Cluster")
    plt.ylabel("Predicción media (normal score)")
    plt.title("Media de predicción por cluster")
    plt.legend()
    plt.grid(True, alpha=0.35)
    plt.tight_layout()
    plt.savefig(IMAGENES_DIR / "resultados_ml_media_por_cluster.png")
    plt.show()

    # Mínimo por cluster
    plt.figure(figsize=(8, 5))
    plt.plot(
        agg_pred_por_cluster.index,
        agg_pred_por_cluster["pred_nscore_knn"]["min"],
        marker="o",
        label="KNN (relleno)",
    )
    plt.plot(
        agg_pred_por_cluster.index,
        agg_pred_por_cluster["pred_nscore_xgb"]["min"],
        marker="o",
        label="XGBoost (relleno)",
    )
    plt.plot(
        agg_real_por_cluster.index,
        agg_real_por_cluster["min"],
        marker="o",
        label="Original",
        linestyle="--",
        color="#7c3aed",
    )
    plt.xlabel("Cluster")
    plt.ylabel("Predicción mínima (normal score)")
    plt.title("Mínimo de predicción por cluster")
    plt.legend()
    plt.grid(True, alpha=0.35)
    plt.tight_layout()
    plt.savefig(IMAGENES_DIR / "resultados_ml_minimo_por_cluster.png")
    plt.show()

    # Máximo por cluster
    plt.figure(figsize=(8, 5))
    plt.plot(
        agg_pred_por_cluster.index,
        agg_pred_por_cluster["pred_nscore_knn"]["max"],
        marker="o",
        label="KNN (relleno)",
    )
    plt.plot(
        agg_pred_por_cluster.index,
        agg_pred_por_cluster["pred_nscore_xgb"]["max"],
        marker="o",
        label="XGBoost (relleno)",
    )
    plt.plot(
        agg_real_por_cluster.index,
        agg_real_por_cluster["max"],
        marker="o",
        label="Original",
        linestyle="--",
        color="#7c3aed",
    )
    plt.xlabel("Cluster")
    plt.ylabel("Predicción máxima (normal score)")
    plt.title("Máximo de predicción por cluster")
    plt.legend()
    plt.grid(True, alpha=0.35)
    plt.tight_layout()
    plt.savefig(IMAGENES_DIR / "resultados_ml_maximo_por_cluster.png")
    plt.show()

# %%
